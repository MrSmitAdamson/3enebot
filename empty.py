from openpyxl import load_workbook
import bot_timing

wb=load_workbook("commands.xlsx")
ws = wb['times']
#today = bot_timing.get_weekday()
today = 'понедельник'
i=1
days_dict = dict()

for com_keys in ws['A']:
    days_list = list()
    for lalka in ws[str(i)]:
        days_list.append(lalka.value)
    days_list.remove(com_keys.value)
    i = i + 1
    days_dict[com_keys.value] = days_list

#i=0
#mes = ''
#while i < len(days_dict[today]):
#    mes = str(days_dict[today][i]) + '\n' + str(mes)
#    i = i + 1

def lessons_time_func(i):
    wb=load_workbook("commands.xlsx")
    ws = wb['times']
    time_list = list()
    for lalka in ws[str(i)]:
        time_list.append(lalka.value)
    print(time_list)

lessons_time_func(5) 



    



