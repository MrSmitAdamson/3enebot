#-*-coding:utf8;-*-
from openpyxl import load_workbook

wb=load_workbook('commands.xlsx')
ws = wb['commands']

def excel_commands_func(mes_text):
    commands_dict = dict()
    i = 1
    for com_keys in ws['A']:
        commands_dict[com_keys.value] = ws['B' + str(i)].value
        i = i + 1
    return (commands_dict[mes_text])

def excel_messages_func(mes_text):
    commands_list = list()
    for com_keys in ws['A']:
        commands_list.append(com_keys.value)
    if mes_text in commands_list:
        return (True)
    else:
        return(False)
