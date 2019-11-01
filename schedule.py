# -*- coding: utf-8 -*-
import bot_timing
from openpyxl import load_workbook

wb=load_workbook("commands.xlsx")
ws = wb['schedule']
i=1
days_dict = dict()
today = bot_timing.get_weekday_func()

for com_keys in ws['A']:
    days_list = list()
    for lessons in ws[str(i)]:
        if str(lessons.value) == 'None':
            days_list.append(' ')
        else:
            days_list.append(lessons.value)
    days_list.remove(com_keys.value)
    i = i + 1
    days_dict[com_keys.value.lower()] = days_list


def get_schedule_func(today):
    if days_dict[today][0].lower() == 'выходной':
        return('Сегодня выходной')
    else:
        return(lessons_func(today))

def lessons_func(day):
    i=0
    today_lessons = f'{day.capitalize()}:'
    while i < len(days_dict[day]):
        today_lessons = str(today_lessons) + '\n' + str(
            lessons_time_func(i+1)[1]) + ':' + str(
            lessons_time_func(i+1)[2]) + ' - ' + str(
            lessons_time_func(i+1)[4]) + ':' + str(
            lessons_time_func(i+1)[5]) + ' ' + str(days_dict[day][i])
        i = i + 1
    return(today_lessons)

def schedule_day(mes_text):
    commands_list = list()
    for com_keys in ws['A']:
        commands_list.append(com_keys.value.lower())
    if mes_text.lower() in commands_list:
        return (True)
    else:
        return(False)
    
def lessons_time_func(i):
    wb=load_workbook("commands.xlsx")
    ws = wb['times']
    time_list = list()
    for lalka in ws[str(i)]:
        if len(str(lalka.value)) == 1:
            time_list.append(f'{0}{lalka.value}')
        else:
            time_list.append(f'{lalka.value}')
    return(time_list)