# -*- coding: utf-8 -*-
import requests
import vk_api
from openpyxl import load_workbook

import bot_report

vk_session = vk_api.VkApi(token='02d23799d5fcd74ab5c28e8ca9b1d268dc43308b34d0918332da3316c60a5762bb467d4c9641a0fc0cdcd')
vk = vk_session.get_api()

wb=load_workbook("commands.xlsx")
ws = wb['users']

def user_group_func():
    users_dict = dict()
    for com_keys in ws['A']:
        for com_keys1 in ws['B']:
            users_dict[com_keys.value] = com_keys1.value


def users_list_get_func():
    users_list = list()
    for com_keys in ws['A']:
        if com_keys.value != None:
            users_list.append(com_keys.value)
    return(users_list)

def user_in_base_func(user_id):
    if user_id in users_list_get_func():
        return True
    else:
        return False

def user_add(user_id, group):
    i = len(users_list_get_func()) + 1
    ws[f'A{i}'] = user_id
    ws[f'B{i}'] = group
    try:
        wb.save("commands.xlsx")
        mes_text = f'Пользователь {vk.users.get(user_id=user_id)[0]["first_name"]} успешно занесен в список'
        bot_report.mes2admin_func(user_id, mes_text)
    except:
        mes_text = f'Пользователя {vk.users.get(user_id=user_id)[0]["first_name"]} не удалось занести в список'
        bot_report.mes2admin_func(user_id, mes_text)